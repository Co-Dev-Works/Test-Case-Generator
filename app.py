# exporter.py
import csv
import io
import json
from typing import List
from models import TestCase
from openpyxl import Workbook
from openpyxl.utils import get_column_letter

def testcases_to_csv_bytes(testcases: List[TestCase]) -> bytes:
    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow(["ID","Title","Steps (newline)","Expected","Priority","Tags","SampleData"])
    for tc in testcases:
        writer.writerow([tc.id, tc.title, "\n".join(tc.steps), tc.expected, tc.priority, ", ".join(tc.tags), json.dumps(tc.sample_data or {})])
    return output.getvalue().encode('utf-8')

def testcases_to_json_bytes(testcases: List[TestCase], meta: dict = None) -> bytes:
    arr = []
    for tc in testcases:
        arr.append({
            "id": tc.id,
            "title": tc.title,
            "steps": tc.steps,
            "expected": tc.expected,
            "priority": tc.priority,
            "tags": tc.tags,
            "sample_data": tc.sample_data or {}
        })
    payload = {"meta": meta or {}, "testcases": arr}
    return json.dumps(payload, indent=2).encode('utf-8')

def testcases_to_xlsx_bytes(testcases: List[TestCase]) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "TestCases"
    headers = ["ID","Title","Steps","Expected","Priority","Tags","SampleData"]
    ws.append(headers)
    for tc in testcases:
        ws.append([tc.id, tc.title, "\n".join(tc.steps), tc.expected, tc.priority, ", ".join(tc.tags), json.dumps(tc.sample_data or {})])
    for i, col in enumerate(ws.columns, start=1):
        max_len = 0
        for cell in col:
            try:
                v = str(cell.value or "")
                if len(v) > max_len:
                    max_len = len(v)
            except:
                pass
        ws.column_dimensions[get_column_letter(i)].width = min(max_len + 2, 50)
    bio = io.BytesIO()
    wb.save(bio)
    bio.seek(0)
    return bio.read()
